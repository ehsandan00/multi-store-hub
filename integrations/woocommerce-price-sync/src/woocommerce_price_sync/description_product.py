from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

DESCRIPTION_HEADERS = [
    "sku",
    "product_name",
    "short_description",
    "full_description",
]

SYNC_RESULT_HEADERS = {
    "woo_product_id",
    "title",
    "code",
    "excel_row",
    "product_type",
    "action",
}


@dataclass
class ProductFacts:
    title: str
    woo_product_id: str | int | None = None
    code: str = ""
    brand: str = ""
    product_type: str = ""
    form: str = ""
    volume: str = ""
    count: str = ""
    main_uses: list[str] = field(default_factory=list)
    benefits: list[tuple[str, str]] = field(default_factory=list)
    ingredients: list[tuple[str, str, str]] = field(default_factory=list)
    usage_steps: list[str] = field(default_factory=list)
    suitable_for: list[str] = field(default_factory=list)
    cautions: list[str] = field(default_factory=list)
    quick_facts: list[tuple[str, str]] = field(default_factory=list)
    faq: list[tuple[str, str]] = field(default_factory=list)
    summary: str = ""
    category_hint: str = ""


@dataclass
class ProductDescription:
    sku: str
    title: str
    short_description: str
    full_description: str
    notes: str = ""


def _slug_words(title: str) -> str:
    cleaned = re.sub(r"\s+", " ", title.strip())
    return cleaned


def generate_seo_title(facts: ProductFacts) -> str:
    parts = [facts.title]
    if facts.brand and facts.brand not in facts.title:
        parts.append(facts.brand)
    if facts.form:
        parts.append(facts.form)
    if facts.volume:
        parts.append(facts.volume)
    return " | ".join(dict.fromkeys(part for part in parts if part))


def generate_short_description(facts: ProductFacts) -> str:
    summary = facts.summary or (
        f"{facts.title} یک انتخاب حرفه‌ای برای مراقبت روزانه است "
        f"و به حفظ ظاهر سالم و متعادل کمک می‌کند."
    )
    uses = facts.main_uses[:3] or ["مراقبت روزانه", "تقویت روتین زیبایی", "حفظ سلامت ظاهری"]
    use_text = "، ".join(uses)
    tips = facts.usage_steps[:2] or [
        "قبل از استفاده، پوست را تمیز و خشک کنید.",
        "محصول را به‌صورت یکنواخت و با مقدار مناسب استفاده کنید.",
    ]
    return (
        f"<p>{summary}</p>"
        f"<p>این محصول برای {use_text} مناسب است. "
        f"{tips[0]} "
        f"{'همچنین ' + tips[1] if len(tips) > 1 else ''}"
        f"در صورت حساسیت پوستی، ابتدا تست پچ انجام دهید.</p>"
    )


def build_infographic_html(facts: ProductFacts) -> str:
    pills = facts.main_uses[:4] or [item[0] for item in facts.benefits[:4]]
    pill_html = "".join(
        f'<span class="info-pill">{_escape_html(item)}</span>' for item in pills if item
    )
    benefit_cards = ""
    for emoji, (title, body) in zip(
        ["✨", "💗", "🌸", "🧴", "💎", "🧬"],
        facts.benefits[:6],
    ):
        benefit_cards += (
            f'<article class="benefit-card">'
            f'<span class="benefit-icon">{emoji}</span>'
            f"<h4>{_escape_html(title)}</h4>"
            f"<p>{_escape_html(body)}</p>"
            f"</article>"
        )
    ingredient_cards = ""
    for name, role, note in facts.ingredients[:6]:
        ingredient_cards += (
            f'<article class="ingredient-card">'
            f"<h4>{_escape_html(name)}</h4>"
            f'<span class="ingredient-role">{_escape_html(role)}</span>'
            f"<p>{_escape_html(note)}</p>"
            f"</article>"
        )
    usage_cards = ""
    for index, step in enumerate(facts.usage_steps[:4], start=1):
        usage_cards += (
            f'<article class="usage-step">'
            f'<span class="step-num">{index}</span>'
            f"<p>{_escape_html(step)}</p>"
            f"</article>"
        )
    quick_facts = ""
    for label, value in facts.quick_facts[:6]:
        quick_facts += (
            f'<div class="fact-item"><span class="fact-label">{_escape_html(label)}</span>'
            f'<strong class="fact-value">{_escape_html(value)}</strong></div>'
        )
    suitable = "".join(f"<li>{_escape_html(item)}</li>" for item in facts.suitable_for[:5])
    cautions = "".join(f"<li>{_escape_html(item)}</li>" for item in facts.cautions[:5])
    takeaway = facts.summary or f"{facts.title} گزینه‌ای مطمئن برای تکمیل روتین مراقبتی شماست."

    return f"""<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
.lux-beauty-infographic,
.lux-beauty-infographic *,
.lux-beauty-infographic *::before,
.lux-beauty-infographic *::after {{
  box-sizing: border-box;
}}
.lux-beauty-infographic {{
  width: 100%;
  max-width: 750px;
  margin: 0 auto;
  color: #5C4E55;
  font-family: inherit;
  line-height: 1.8;
}}
.lux-beauty-infographic .hero {{
  background: linear-gradient(145deg, #FFFFFF 0%, #F7F0F3 55%, #FFF9FB 100%);
  border: 1px solid #D9C6CC;
  border-radius: 28px;
  padding: clamp(18px, 4vw, 28px);
  box-shadow: 0 10px 28px rgba(166, 130, 151, 0.12);
  margin-bottom: 18px;
}}
.lux-beauty-infographic .hero h2 {{
  margin: 0 0 10px;
  color: #8C6A7C;
  font-size: clamp(18px, 4.5vw, 22px);
  line-height: 1.5;
}}
.lux-beauty-infographic .hero p {{
  margin: 0 0 14px;
  color: #7A6F75;
  font-size: clamp(14px, 3.6vw, 15px);
}}
.lux-beauty-infographic .pill-row {{
  display: flex;
  flex-wrap: wrap;
  gap: 8px;
}}
.lux-beauty-infographic .info-pill {{
  display: inline-flex;
  align-items: center;
  padding: 6px 12px;
  border-radius: 999px;
  background: #D1B3C4;
  color: #5C4E55;
  font-size: 13px;
}}
.lux-beauty-infographic .section {{
  margin-bottom: 16px;
}}
.lux-beauty-infographic .section h3 {{
  margin: 0 0 12px;
  color: #A68297;
  font-size: clamp(16px, 4vw, 19px);
}}
.lux-beauty-infographic .facts-grid {{
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
  gap: 10px;
}}
.lux-beauty-infographic .fact-item {{
  background: #FFFFFF;
  border: 1px solid #D9C6CC;
  border-radius: 16px;
  padding: 12px;
}}
.lux-beauty-infographic .fact-label {{
  display: block;
  color: #7A6F75;
  font-size: 13px;
  margin-bottom: 4px;
}}
.lux-beauty-infographic .fact-value {{
  color: #5C4E55;
  font-size: 14px;
}}
.lux-beauty-infographic .benefit-grid {{
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
  gap: 12px;
}}
.lux-beauty-infographic .benefit-card,
.lux-beauty-infographic .ingredient-card,
.lux-beauty-infographic .usage-step {{
  background: #F7F0F3;
  border: 1px solid #E6D5DF;
  border-radius: 18px;
  padding: 14px;
}}
.lux-beauty-infographic .benefit-icon {{
  display: inline-flex;
  width: 34px;
  height: 34px;
  align-items: center;
  justify-content: center;
  border-radius: 50%;
  background: #FFFFFF;
  margin-bottom: 8px;
}}
.lux-beauty-infographic .benefit-card h4,
.lux-beauty-infographic .ingredient-card h4 {{
  margin: 0 0 6px;
  color: #8C6A7C;
  font-size: clamp(15px, 3.8vw, 17px);
}}
.lux-beauty-infographic .benefit-card p,
.lux-beauty-infographic .ingredient-card p,
.lux-beauty-infographic .usage-step p {{
  margin: 0;
  color: #7A6F75;
  font-size: 14px;
}}
.lux-beauty-infographic .ingredient-role {{
  display: inline-block;
  margin-bottom: 8px;
  padding: 4px 10px;
  border-radius: 999px;
  background: #E6D5DF;
  color: #5C4E55;
  font-size: 13px;
}}
.lux-beauty-infographic .usage-grid {{
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
  gap: 10px;
}}
.lux-beauty-infographic .usage-step {{
  display: flex;
  gap: 10px;
  align-items: flex-start;
}}
.lux-beauty-infographic .step-num {{
  flex: 0 0 28px;
  width: 28px;
  height: 28px;
  border-radius: 50%;
  background: #A68297;
  color: #FFFFFF;
  display: inline-flex;
  align-items: center;
  justify-content: center;
  font-size: 13px;
}}
.lux-beauty-infographic .split-grid {{
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
  gap: 12px;
}}
.lux-beauty-infographic .panel {{
  background: #FFFFFF;
  border: 1px solid #D9C6CC;
  border-radius: 18px;
  padding: 14px;
}}
.lux-beauty-infographic .panel h4 {{
  margin: 0 0 8px;
  color: #8C6A7C;
  font-size: 15px;
}}
.lux-beauty-infographic .panel ul {{
  margin: 0;
  padding: 0 18px 0 0;
  color: #7A6F75;
  font-size: 14px;
}}
.lux-beauty-infographic .takeaway {{
  background: linear-gradient(90deg, #F7F0F3 0%, #FFF9FB 100%);
  border: 1px solid #D1B3C4;
  border-radius: 20px;
  padding: 14px 16px;
  color: #5C4E55;
  font-size: 14px;
}}
@media (max-width: 600px) {{
  .lux-beauty-infographic .hero,
  .lux-beauty-infographic .panel,
  .lux-beauty-infographic .benefit-card,
  .lux-beauty-infographic .ingredient-card,
  .lux-beauty-infographic .usage-step,
  .lux-beauty-infographic .fact-item {{
    padding: 12px;
  }}
}}
@media (max-width: 400px) {{
  .lux-beauty-infographic .benefit-grid,
  .lux-beauty-infographic .usage-grid {{
    grid-template-columns: 1fr;
  }}
}}
</style>
</head>
<body>
<div class="lux-beauty-infographic">
  <section class="hero">
    <h2>{_escape_html(facts.title)}</h2>
    <p>{_escape_html(takeaway)}</p>
    <div class="pill-row">{pill_html}</div>
  </section>
  <section class="section">
    <h3>اطلاعات سریع محصول</h3>
    <div class="facts-grid">{quick_facts}</div>
  </section>
  <section class="section">
    <h3>مزایای کلیدی</h3>
    <div class="benefit-grid">{benefit_cards}</div>
  </section>
  <section class="section">
    <h3>ترکیبات و فرمول فعال</h3>
    <div class="benefit-grid">{ingredient_cards}</div>
  </section>
  <section class="section">
    <h3>روش استفاده</h3>
    <div class="usage-grid">{usage_cards}</div>
  </section>
  <section class="section split-grid">
    <article class="panel">
      <h4>مناسب برای</h4>
      <ul>{suitable}</ul>
    </article>
    <article class="panel">
      <h4>نکات مهم</h4>
      <ul>{cautions}</ul>
    </article>
  </section>
  <section class="takeaway">{_escape_html(takeaway)}</section>
</div>
</body>
</html>"""


def product_sku(facts: ProductFacts) -> str:
    for value in (facts.code, facts.woo_product_id):
        if value not in (None, ""):
            return str(value)
    return ""


def build_product_table_rows(facts: ProductFacts) -> list[tuple[str, str, str]]:
    rows: list[tuple[str, str, str]] = []
    seen: set[str] = set()

    def add(label: str, value: str, note: str = "—") -> None:
        key = label.casefold()
        if not value or key in seen:
            return
        seen.add(key)
        rows.append((label, value, note))

    add("نام محصول", facts.title)
    if facts.brand:
        add("برند", facts.brand)
    if facts.product_type:
        add("نوع محصول", facts.product_type)
    if facts.form:
        add("فرم", facts.form)
    if facts.volume:
        add("حجم / اندازه", facts.volume)
    if facts.count:
        add("تعداد", facts.count)
    for label, value in facts.quick_facts:
        add(label, value)
    for name, amount, nrv in facts.ingredients:
        add(name, amount, nrv)
    for title, body in facts.benefits[:4]:
        add(f"مزیت: {title}", body)
    if facts.usage_steps:
        add("روش مصرف", "؛ ".join(facts.usage_steps[:3]))
    if facts.suitable_for:
        add("مناسب برای", "، ".join(facts.suitable_for[:4]))
    if facts.cautions:
        add("نکات مهم", "؛ ".join(facts.cautions[:3]))
    return rows[:12]


def build_product_faq_list(facts: ProductFacts) -> list[tuple[str, str]]:
    faq: list[tuple[str, str]] = []
    seen_questions: set[str] = set()

    def add(question: str, answer: str) -> None:
        key = question.casefold()
        if key in seen_questions or not answer:
            return
        seen_questions.add(key)
        faq.append((question, answer))

    for question, answer in facts.faq:
        add(question, answer)

    if facts.usage_steps:
        add(
            f"روش استفاده از {facts.title} چگونه است؟",
            " ".join(facts.usage_steps[:3]),
        )
    if facts.brand:
        add(
            f"برند {facts.title} چیست؟",
            f"{facts.title} محصول برند {facts.brand} است.",
        )
    if facts.product_type:
        add(
            f"{facts.title} برای چه منظوری است؟",
            f"این محصول در دسته {facts.product_type} قرار دارد "
            f"و برای {('، '.join(facts.main_uses[:3]) if facts.main_uses else 'مراقبت روزانه')} مناسب است.",
        )
    if facts.suitable_for:
        add(
            f"{facts.title} برای چه کسانی مناسب است؟",
            "، ".join(facts.suitable_for[:4]),
        )
    if facts.cautions:
        add(
            f"نکات مهم هنگام استفاده از {facts.title} چیست؟",
            "؛ ".join(facts.cautions[:3]),
        )
    if facts.benefits:
        title, body = facts.benefits[0]
        add(
            f"مهم‌ترین مزیت {facts.title} چیست؟",
            f"{title}: {body}",
        )
    if facts.ingredients and facts.product_type == "مکمل غذایی":
        names = "، ".join(name for name, _, _ in facts.ingredients[:4])
        add(
            f"ترکیبات اصلی {facts.title} کدامند؟",
            f"فرمول شامل {names} است.",
        )
    elif facts.ingredients and facts.product_type == "عطر":
        names = "، ".join(name for name, role, _ in facts.ingredients[:4])
        add(
            f"نت‌های بویایی {facts.title} چیست؟",
            f"ترکیب رایحه شامل {names} است.",
        )
    elif facts.ingredients:
        names = "، ".join(name for name, role, _ in facts.ingredients[:4])
        add(
            f"ترکیبات کلیدی {facts.title} چیست؟",
            f"فرمول حاوی {names} است.",
        )

    add(
        f"آیا {facts.title} اصل است؟",
        "محصول را از فروشگاه معتبر با ضمانت اصالت تهیه کنید.",
    )
    return faq[:8]


def build_table_html(facts: ProductFacts) -> str:
    table_rows = build_product_table_rows(facts)
    is_supplement = facts.product_type == "مکمل غذایی" and bool(facts.ingredients)
    if is_supplement and facts.ingredients:
        headers = ("مشخصه / ترکیب", "مقدار", "درصد نیاز روزانه")
    else:
        headers = ("مشخصه", "جزئیات", "توضیح")

    rows = ""
    for name, amount, nrv in table_rows:
        rows += (
            "<tr>"
            f"<td>{_escape_html(name)}</td>"
            f"<td>{_escape_html(amount)}</td>"
            f"<td>{_escape_html(nrv)}</td>"
            "</tr>"
        )
    return f"""<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
.rose-table-block {{
  width: 100%;
  max-width: 750px;
  margin: 0 auto;
  color: #5C4E55;
  font-family: inherit;
  direction: rtl;
}}
.rose-table-block .table-card {{
  background: #FFFFFF;
  border: 1px solid #D9C6CC;
  border-radius: 18px;
  box-shadow: 0 8px 24px rgba(166, 130, 151, 0.1);
  overflow: hidden;
}}
.rose-table-block .table-scroll {{
  overflow-x: auto;
  -webkit-overflow-scrolling: touch;
}}
.rose-table-block .table-scroll::-webkit-scrollbar {{
  height: 8px;
}}
.rose-table-block .table-scroll::-webkit-scrollbar-track {{
  background: #F7F0F3;
}}
.rose-table-block .table-scroll::-webkit-scrollbar-thumb {{
  background: #A68297;
  border-radius: 999px;
}}
.rose-table-block .rose-table {{
  width: 100%;
  border-collapse: collapse;
  min-width: 520px;
  font-size: 14px;
}}
.rose-table-block .rose-table th {{
  background: #E6D5DF;
  color: #5C4E55;
  text-align: right;
  padding: 12px 14px;
  border-bottom: 1px solid #D9C6CC;
}}
.rose-table-block .rose-table td {{
  text-align: right;
  padding: 12px 14px;
  border-bottom: 1px solid #D9C6CC;
  color: #7A6F75;
  vertical-align: top;
}}
.rose-table-block .rose-table tbody tr:nth-child(even) td {{
  background: #FFF9FB;
}}
.rose-table-block .rose-table tbody tr:hover td {{
  background: #F7F0F3;
}}
@media (max-width: 600px) {{
  .rose-table-block .rose-table {{
    font-size: 13px;
  }}
  .rose-table-block .rose-table th,
  .rose-table-block .rose-table td {{
    padding: 10px 12px;
  }}
}}
</style>
</head>
<body>
<div class="rose-table-block">
  <section class="table-card">
    <div class="table-scroll" aria-label="جدول قابل اسکرول">
      <table class="rose-table">
        <thead>
          <tr>
            <th>{_escape_html(headers[0])}</th>
            <th>{_escape_html(headers[1])}</th>
            <th>{_escape_html(headers[2])}</th>
          </tr>
        </thead>
        <tbody>
          {rows}
        </tbody>
      </table>
    </div>
  </section>
</div>
</body>
</html>"""


def build_faq_html(facts: ProductFacts) -> str:
    items = ""
    for question, answer in build_product_faq_list(facts):
        items += (
            '<article class="faq-item">'
            f'<button class="faq-question" type="button" aria-expanded="false">'
            f"<span>{_escape_html(question)}</span>"
            '<span class="faq-icon">+</span>'
            "</button>"
            f'<div class="faq-answer"><p>{_escape_html(answer)}</p></div>'
            "</article>"
        )
    return f"""<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
.faq-lilac-block {{
  width: 100%;
  max-width: 750px;
  margin: 0 auto;
  color: #5C4E55;
  font-family: inherit;
  direction: rtl;
}}
.faq-lilac-block .faq-card {{
  background: #F7F0F3;
  border: 1px solid #D9C6CC;
  border-radius: 20px;
  padding: clamp(14px, 3vw, 20px);
  box-shadow: 0 8px 22px rgba(166, 130, 151, 0.1);
}}
.faq-lilac-block .faq-title {{
  margin: 0 0 14px;
  color: #A68297;
  font-size: clamp(16px, 4vw, 18px);
}}
.faq-lilac-block .faq-item {{
  background: #FFFFFF;
  border: 1px solid #D9C6CC;
  border-radius: 14px;
  overflow: hidden;
  margin-bottom: 10px;
}}
.faq-lilac-block .faq-question {{
  width: 100%;
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 12px;
  padding: 14px 16px;
  background: transparent;
  border: 0;
  color: #5C4E55;
  font: inherit;
  text-align: right;
  cursor: pointer;
}}
.faq-lilac-block .faq-question:hover {{
  background: #FFF9FB;
}}
.faq-lilac-block .faq-icon {{
  color: #A68297;
  font-size: 18px;
  line-height: 1;
}}
.faq-lilac-block .faq-answer {{
  max-height: 0;
  overflow: hidden;
  transition: max-height 0.25s ease;
}}
.faq-lilac-block .faq-item.open .faq-answer {{
  max-height: 320px;
}}
.faq-lilac-block .faq-item.open .faq-icon {{
  transform: rotate(45deg);
}}
.faq-lilac-block .faq-answer p {{
  margin: 0;
  padding: 0 16px 14px;
  color: #7A6F75;
  font-size: 14px;
  line-height: 1.9;
}}
</style>
</head>
<body>
<div class="faq-lilac-block">
  <section class="faq-card">
    <h3 class="faq-title">سوالات متداول</h3>
    {items}
  </section>
</div>
<script>
document.querySelectorAll('.faq-lilac-block .faq-question').forEach(function (button) {{
  button.addEventListener('click', function () {{
    var item = button.closest('.faq-item');
    var open = item.classList.contains('open');
    document.querySelectorAll('.faq-lilac-block .faq-item.open').forEach(function (el) {{
      el.classList.remove('open');
      el.querySelector('.faq-question').setAttribute('aria-expanded', 'false');
    }});
    if (!open) {{
      item.classList.add('open');
      button.setAttribute('aria-expanded', 'true');
    }}
  }});
}});
</script>
</body>
</html>"""


def generate_long_description(facts: ProductFacts) -> str:
    seo_title = generate_seo_title(facts)
    intro = facts.summary or (
        f"{facts.title} با فرمولی متعادل و حرفه‌ای، انتخابی مناسب برای تکمیل "
        "روتین مراقبتی روزانه شماست."
    )
    benefit_lines = "".join(
        f"<li><strong>{_escape_html(title)}:</strong> {_escape_html(body)}</li>"
        for title, body in facts.benefits[:5]
    )
    geo_line = ""
    if facts.brand:
        geo_line = (
            f"<p>خرید آنلاین <strong>{_escape_html(facts.title)}</strong> "
            f"اصل برند <strong>{_escape_html(facts.brand)}</strong> "
            f"با ارسال سریع در سراسر ایران.</p>"
        )
    return (
        f"<h2>{_escape_html(seo_title)}</h2>"
        f"<p>{_escape_html(intro)}</p>"
        f"{geo_line}"
        f"<h3>مزایای اصلی</h3><ul>{benefit_lines}</ul>"
        f"<h3>راهنمای استفاده</h3>"
        f"<ol>{''.join(f'<li>{_escape_html(step)}</li>' for step in facts.usage_steps[:4])}</ol>"
    )


def build_full_description(description: str, table_html: str, faq_html: str) -> str:
    return f"{description}\n{table_html}\n{faq_html}"


def build_description(facts: ProductFacts) -> ProductDescription:
    description = generate_long_description(facts)
    table_html = build_table_html(facts)
    faq_html = build_faq_html(facts)
    return ProductDescription(
        sku=product_sku(facts),
        title=facts.title,
        short_description=generate_short_description(facts),
        full_description=build_full_description(description, table_html, faq_html),
    )


def _escape_html(value: str) -> str:
    return (
        str(value)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _header_map(headers: tuple[Any, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, header in enumerate(headers):
        key = _cell_text(header).casefold()
        if key:
            mapping[key] = index
    return mapping


def _read_parent_products_from_rows(rows: list[tuple[Any, ...]]) -> list[ProductFacts]:
    if not rows:
        return []
    headers = _header_map(rows[0])
    products: list[ProductFacts] = []
    seen: set[str] = set()
    for row in rows[1:]:
        title = _pick(row, headers, {"title", "نام کالا"})
        if not title:
            continue
        product_type = _pick(row, headers, {"product_type"})
        if product_type == "variation":
            continue
        attr_name = _pick(row, headers, {"attribute_name", "تنوع"})
        attr_value = _pick(row, headers, {"attribute_value", "مقدار تنوع"})
        if attr_name and attr_value:
            continue
        woo_id = _pick(row, headers, {"woo_product_id"})
        code = _pick(row, headers, {"code", "کد کالا"})
        dedupe_key = f"{code}:{title}" if code else title
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        products.append(
            ProductFacts(
                title=title,
                woo_product_id=woo_id or None,
                code=code,
                product_type=product_type,
            )
        )
    return products


def read_sync_report_products(path: str | Path) -> list[ProductFacts]:
    source = Path(path)
    if not source.is_file():
        raise FileNotFoundError(f"Sync report not found: {source}")

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        if "Results" in workbook.sheetnames:
            sheet = workbook["Results"]
        else:
            sheet = workbook.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))
        return _read_parent_products_from_rows(rows)
    finally:
        workbook.close()


def read_catalog_products(path: str | Path) -> list[ProductFacts]:
    source = Path(path)
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))
        return _read_parent_products_from_rows(rows)
    finally:
        workbook.close()


def _pick(row: tuple[Any, ...], headers: dict[str, int], aliases: set[str]) -> str:
    for alias in aliases:
        index = headers.get(alias.casefold())
        if index is None or index >= len(row):
            continue
        value = _cell_text(row[index])
        if value:
            return value
    return ""


def read_descriptions_report(path: str | Path) -> dict[str, ProductDescription]:
    source = Path(path)
    if not source.is_file():
        return {}
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return {}
        headers = _header_map(rows[0])
        results: dict[str, ProductDescription] = {}
        for row in rows[1:]:
            title = _pick(row, headers, {"product_name", "title", "نام کالا"})
            if not title:
                continue
            sku = _pick(row, headers, {"sku", "code", "کد کالا"})
            short_description = _pick(row, headers, {"short_description"})
            full_description = _pick(row, headers, {"full_description"})
            if not full_description:
                description = _pick(row, headers, {"description"})
                table_html = _pick(row, headers, {"table_html"})
                faq_html = _pick(row, headers, {"faq_html"})
                if description or table_html or faq_html:
                    full_description = build_full_description(description, table_html, faq_html)
            key = f"{sku}:{title}" if sku else title
            results[key] = ProductDescription(
                sku=sku,
                title=title,
                short_description=short_description,
                full_description=full_description,
                notes=_pick(row, headers, {"notes"}),
            )
        return results
    finally:
        workbook.close()


def write_descriptions_report(
    path: str | Path,
    descriptions: list[ProductDescription],
    *,
    input_path: str | Path | None = None,
) -> Path:
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Descriptions"
    sheet.sheet_view.rightToLeft = True
    sheet.freeze_panes = "A2"
    sheet.append(DESCRIPTION_HEADERS)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for item in descriptions:
        sheet.append(
            [
                item.sku,
                item.title,
                item.short_description,
                item.full_description,
            ]
        )

    widths = [18, 42, 40, 80]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width

    if input_path:
        summary = workbook.create_sheet("Summary")
        summary.append(["metric", "value"])
        summary.append(["input", str(input_path)])
        summary.append(["output", str(destination)])
        summary.append(["total_descriptions", len(descriptions)])

    workbook.save(destination)
    return destination


def merge_descriptions(
    existing: dict[str, ProductDescription],
    new_items: list[ProductDescription],
) -> list[ProductDescription]:
    merged = dict(existing)
    for item in new_items:
        key = f"{item.sku}:{item.title}" if item.sku else item.title
        merged[key] = item
    return sorted(merged.values(), key=lambda row: _slug_words(row.title))


def description_key(item: ProductDescription) -> str:
    return f"{item.sku}:{item.title}" if item.sku else item.title


def default_paths(base_dir: Path | None = None) -> tuple[Path, Path, Path]:
    root = base_dir or Path(__file__).resolve().parents[2] / "data"
    return (
        root / "sync-report.xlsx",
        root / "catalog.xlsx",
        root / "descriptions-report.xlsx",
    )


def load_pending_products(
    *,
    catalog: Path,
    descriptions_report: Path,
    limit: int | None = None,
    refresh_fallbacks: bool = False,
    force: bool = False,
) -> list[ProductFacts]:
    if not catalog.is_file():
        raise FileNotFoundError(f"Catalog not found: {catalog}")

    products = read_catalog_products(catalog)
    existing = read_descriptions_report(descriptions_report) if descriptions_report.is_file() else {}
    pending: list[ProductFacts] = []
    for product in products:
        sku = product_sku(product)
        key = f"{sku}:{product.title}" if sku else product.title
        current = existing.get(key)
        if current and not force and not (refresh_fallbacks and current.notes in {"category-fallback", ""}):
            continue
        pending.append(product)
    if limit is not None:
        pending = pending[:limit]
    return pending


def build_parser() -> argparse.ArgumentParser:
    _, catalog, descriptions_report = default_paths()
    parser = argparse.ArgumentParser(
        description="List pending Persian product descriptions from catalog.xlsx.",
    )
    parser.add_argument(
        "--catalog",
        type=Path,
        default=catalog,
        help="Path to catalog.xlsx (parent + simple products only)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=descriptions_report,
        help="Path for descriptions-report.xlsx",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=20,
        help="Maximum number of pending products to list for generation",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Regenerate all products regardless of existing output",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    pending = load_pending_products(
        catalog=args.catalog,
        descriptions_report=args.output,
        limit=args.limit,
        force=args.force,
    )
    if not pending:
        print("No pending products found.")
        return 0
    print(f"Pending products: {len(pending)}")
    for product in pending:
        label = product.woo_product_id or product.code or "-"
        print(f"- {product.title} ({label})")
    print(f"Output report: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
