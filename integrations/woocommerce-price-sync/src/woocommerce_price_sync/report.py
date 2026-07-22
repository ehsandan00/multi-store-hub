from __future__ import annotations

from collections import Counter
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .models import SyncResult, SyncSummary

RESULT_HEADERS = [
    "excel_row",
    "code",
    "title",
    "product_type",
    "attribute_name",
    "attribute_value",
    "match_status",
    "action",
    "woo_product_id",
    "woo_variation_id",
    "old_price",
    "new_price",
    "old_stock_status",
    "new_stock_status",
    "reason",
]


def summarize(results: list[SyncResult], *, apply: bool) -> SyncSummary:
    counts = Counter(result.action for result in results)
    return SyncSummary(
        mode="apply" if apply else "dry-run",
        total_rows=len(results),
        action_counts=dict(sorted(counts.items())),
    )


def default_report_path(input_path: str | Path) -> Path:
    source = Path(input_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return source.with_name(f"{source.stem}_sync_report_{timestamp}.xlsx")


def write_report(
    path: str | Path,
    results: list[SyncResult],
    *,
    apply: bool,
    input_path: str | Path,
) -> Path:
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    summary = summarize(results, apply=apply)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    generated_at = datetime.now(UTC).isoformat()
    summary_rows = [
        ("Mode", summary.mode),
        ("Input file", str(Path(input_path))),
        ("Generated at UTC", generated_at),
        ("Total result rows", summary.total_rows),
    ]
    for action, count in summary.action_counts.items():
        summary_rows.append((f"Action: {action}", count))
    for row in summary_rows:
        summary_sheet.append(row)
    summary_sheet.column_dimensions["A"].width = 24
    summary_sheet.column_dimensions["B"].width = 90
    summary_sheet["A1"].font = Font(bold=True)

    results_sheet = workbook.create_sheet("Results")
    results_sheet.sheet_view.rightToLeft = True
    results_sheet.freeze_panes = "A2"
    results_sheet.auto_filter.ref = f"A1:O{max(len(results) + 1, 2)}"
    results_sheet.append(RESULT_HEADERS)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in results_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for result in results:
        results_sheet.append(
            [
                result.excel_row,
                _excel_safe(result.code),
                _excel_safe(result.title),
                result.product_type,
                _excel_safe(result.attribute_name),
                _excel_safe(result.attribute_value),
                result.match_status,
                result.action,
                result.woo_product_id,
                result.woo_variation_id,
                result.old_price,
                result.new_price,
                result.old_stock_status,
                result.new_stock_status,
                _excel_safe(result.reason),
            ]
        )

    widths = [12, 18, 48, 14, 20, 32, 16, 16, 18, 19, 16, 16, 16, 16, 70]
    for index, width in enumerate(widths, start=1):
        results_sheet.column_dimensions[chr(64 + index)].width = width

    workbook.save(destination)
    return destination


def _excel_safe(value: str) -> str:
    if value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value
