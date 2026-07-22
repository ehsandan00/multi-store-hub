from __future__ import annotations

import argparse
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from .excel import HEADER_ALIASES, REQUIRED_HEADERS, _cell_text, _has_variation_attrs, read_products

PC_CODE_PATTERN = re.compile(r"^PC-(\d+)$", re.IGNORECASE)


@dataclass
class RowRecord:
    row_number: int
    values: list[Any]
    code: str = ""
    reference: str = ""
    title: str = ""
    attribute_name: str = ""
    attribute_value: str = ""
    price: Any = None

    @property
    def has_attrs(self) -> bool:
        return _has_variation_attrs(self.attribute_name, self.attribute_value)

    @property
    def is_empty_duplicate(self) -> bool:
        return not self.has_attrs and not self.reference and self.price in (None, "")


@dataclass
class FixChange:
    row: int
    field: str
    old: str
    new: str
    reason: str


@dataclass
class FixReport:
    input_path: str
    output_path: str
    original_rows: int = 0
    output_rows: int = 0
    removed_rows: int = 0
    code_fixes: int = 0
    reference_fixes: int = 0
    sku_reassignments: int = 0
    changes: list[FixChange] = field(default_factory=list)
    removed: list[tuple[int, str, str]] = field(default_factory=list)


def _map_header_indexes(headers: tuple[Any, ...]) -> dict[str, int]:
    indexes: dict[str, int] = {}
    for index, value in enumerate(headers):
        header = _cell_text(value).casefold()
        for field, aliases in HEADER_ALIASES.items():
            if header in {alias.casefold() for alias in aliases}:
                indexes.setdefault(field, index)
                break
    missing = sorted(REQUIRED_HEADERS - indexes.keys())
    if missing:
        raise ValueError(f"Missing required Excel columns: {', '.join(missing)}")
    return indexes


def _record_from_row(row_number: int, values: tuple[Any, ...], indexes: dict[str, int]) -> RowRecord:
    row_values = list(values)
    return RowRecord(
        row_number=row_number,
        values=row_values,
        code=_cell_text(row_values[indexes["code"]] if indexes["code"] < len(row_values) else None),
        reference=_cell_text(row_values[indexes["reference"]] if indexes["reference"] < len(row_values) else None),
        title=_cell_text(row_values[indexes["title"]] if indexes["title"] < len(row_values) else None),
        attribute_name=_cell_text(
            row_values[indexes["attribute_name"]] if indexes["attribute_name"] < len(row_values) else None
        ),
        attribute_value=_cell_text(
            row_values[indexes["attribute_value"]] if indexes["attribute_value"] < len(row_values) else None
        ),
        price=row_values[indexes["price"]] if indexes["price"] < len(row_values) else None,
    )


def _set_field(record: RowRecord, indexes: dict[str, int], field: str, value: Any) -> None:
    index = indexes[field]
    while len(record.values) <= index:
        record.values.append(None)
    record.values[index] = value
    if field == "code":
        record.code = _cell_text(value)
    elif field == "reference":
        record.reference = _cell_text(value)
    elif field == "title":
        record.title = _cell_text(value)
    elif field == "attribute_name":
        record.attribute_name = _cell_text(value)
    elif field == "attribute_value":
        record.attribute_value = _cell_text(value)
    elif field == "price":
        record.price = value


def _next_pc_code(used: set[str]) -> str:
    max_number = 0
    for code in used:
        match = PC_CODE_PATTERN.match(code)
        if match:
            max_number = max(max_number, int(match.group(1)))
    while True:
        max_number += 1
        candidate = f"PC-{max_number}"
        if candidate not in used:
            used.add(candidate)
            return candidate


def _load_records(path: Path) -> tuple[list[Any], dict[str, int], list[RowRecord]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        row_iter = worksheet.iter_rows(values_only=True)
        headers = next(row_iter)
        indexes = _map_header_indexes(headers)
        records: list[RowRecord] = []
        for row_number, values in enumerate(row_iter, start=2):
            if not any(_cell_text(value) for value in values):
                continue
            records.append(_record_from_row(row_number, values, indexes))
        return list(headers), indexes, records
    finally:
        workbook.close()


def fix_workbook_records(
    records: list[RowRecord],
    indexes: dict[str, int],
) -> tuple[FixReport, list[RowRecord]]:
    report = FixReport(input_path="", output_path="")
    report.original_rows = len(records)

    used_codes = {record.code for record in records if record.code}

    # 1) Move misplaced SKUs from reference column into code column.
    for record in records:
        if record.code:
            continue
        if not record.reference or not PC_CODE_PATTERN.match(record.reference):
            continue
        if record.has_attrs:
            continue
        new_code = record.reference
        report.changes.append(
            FixChange(
                record.row_number,
                "code",
                "",
                new_code,
                "SKU was in parent column; moved to code column",
            )
        )
        _set_field(record, indexes, "code", new_code)
        _set_field(record, indexes, "reference", "")
        record.code = new_code
        record.reference = ""
        used_codes.add(new_code)
        report.code_fixes += 1

    # 2) Remove junk duplicate rows and duplicate parent placeholders.
    parent_refs = {
        record.reference
        for record in records
        if record.reference and record.has_attrs
    }
    by_code: dict[str, list[RowRecord]] = defaultdict(list)
    for record in records:
        if record.code:
            by_code[record.code].append(record)

    kept: list[RowRecord] = []
    removed_rows: set[int] = set()
    for code, group in by_code.items():
        if len(group) == 1:
            kept.append(group[0])
            continue

        group = sorted(group, key=lambda row: row.row_number)
        parents = [row for row in group if code in parent_refs and not row.has_attrs]
        variations = [row for row in group if row.has_attrs]
        others = [
            row
            for row in group
            if row not in parents
            and row not in variations
            and not (row.is_empty_duplicate and (variations or parents))
        ]

        survivors = list(variations) + list(others)
        if parents:
            survivors.insert(0, parents[0])

        for row in group:
            if row in survivors:
                continue
            removed_rows.add(row.row_number)
            report.removed.append((row.row_number, code, row.title))
            if row in parents[1:]:
                reason = "Removed duplicate parent placeholder row"
            elif row.is_empty_duplicate:
                reason = "Removed empty duplicate row for this SKU"
            else:
                reason = "Removed duplicate row for this SKU"
            report.changes.append(
                FixChange(row.row_number, "row", "kept", "removed", reason)
            )

        kept.extend(survivors)

    for record in records:
        if not record.code and record.row_number not in removed_rows:
            kept.append(record)

    fixed_records = sorted(kept, key=lambda row: row.row_number)
    report.removed_rows = len(removed_rows)

    # 3) Resolve remaining SKU conflicts.
    used_codes = {record.code for record in fixed_records if record.code}
    parent_refs = {
        record.reference
        for record in fixed_records
        if record.reference and record.has_attrs
    }
    by_code = defaultdict(list)
    for record in fixed_records:
        if record.code:
            by_code[record.code].append(record)

    for code, group in by_code.items():
        if len(group) <= 1:
            continue

        variations = [row for row in group if row.has_attrs]
        parents = [row for row in group if not row.has_attrs and code in parent_refs]
        simple_rows = [row for row in group if row not in variations and row not in parents]

        if parents:
            keeper = sorted(parents, key=lambda row: row.row_number)[0]
            reassign = [row for row in group if row is not keeper]
            reason = "Parent product keeps the original SKU"
        elif variations:
            keeper = sorted(variations, key=lambda row: row.row_number)[0]
            reassign = [row for row in group if row is not keeper]
            reason = "Earliest variation keeps the original SKU"
        else:
            keeper = sorted(simple_rows or group, key=lambda row: row.row_number)[0]
            reassign = [row for row in group if row is not keeper]
            reason = "Earliest row keeps the original SKU"

        for row in reassign:
            new_code = _next_pc_code(used_codes)
            report.changes.append(
                FixChange(row.row_number, "code", code, new_code, reason)
            )
            _set_field(row, indexes, "code", new_code)
            row.code = new_code
            report.sku_reassignments += 1

    # 4) Parent rows should not carry a parent reference.
    for record in fixed_records:
        if record.code in parent_refs and not record.has_attrs and record.reference:
            report.changes.append(
                FixChange(
                    record.row_number,
                    "reference",
                    record.reference,
                    "",
                    "Parent product row should not reference another parent",
                )
            )
            _set_field(record, indexes, "reference", "")
            record.reference = ""
            report.reference_fixes += 1

    report.output_rows = len(fixed_records)
    return report, fixed_records


def write_fixed_workbook(
    headers: list[Any],
    records: list[RowRecord],
    output_path: Path,
    report: FixReport,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(headers)
    for record in records:
        sheet.append(record.values)

    changes_sheet = workbook.create_sheet("FixLog")
    changes_sheet.append(["row", "field", "old", "new", "reason"])
    for change in report.changes:
        changes_sheet.append([change.row, change.field, change.old, change.new, change.reason])

    summary = workbook.create_sheet("Summary")
    summary.append(["metric", "value"])
    for key, value in [
        ("input", report.input_path),
        ("output", report.output_path),
        ("original_rows", report.original_rows),
        ("output_rows", report.output_rows),
        ("removed_rows", report.removed_rows),
        ("code_fixes", report.code_fixes),
        ("reference_fixes", report.reference_fixes),
        ("sku_reassignments", report.sku_reassignments),
    ]:
        summary.append([key, value])

    for ws in (changes_sheet, summary):
        header = ws[1]
        for cell in header:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")

    workbook.save(output_path)


def fix_excel_file(input_path: str | Path, output_path: str | Path | None = None) -> FixReport:
    source = Path(input_path)
    if not source.is_file():
        raise FileNotFoundError(f"Excel file not found: {source}")
    destination = Path(output_path) if output_path else source.with_name(f"{source.stem}_cleaned.xlsx")

    headers, indexes, records = _load_records(source)
    report, records = fix_workbook_records(records, indexes)
    report.input_path = str(source)
    report.output_path = str(destination)
    write_fixed_workbook(headers, records, destination, report)
    return report


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Clean and normalize product Excel workbooks.")
    parser.add_argument("input", type=Path, help="Source .xlsx workbook")
    parser.add_argument("-o", "--output", type=Path, help="Output .xlsx path")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        report = fix_excel_file(args.input, args.output)
        validation = read_products(report.output_path)
    except (FileNotFoundError, ValueError) as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 2

    issue_counts = Counter(issue.reason for issue in validation.issues)
    duplicate_codes = len(
        {code for code, count in Counter(product.code for product in validation.products).items() if count > 1}
    )
    print(f"Output: {report.output_path}")
    print(f"Rows: {report.original_rows} -> {report.output_rows} (removed {report.removed_rows})")
    print(f"Code fixes: {report.code_fixes}")
    print(f"Reference fixes: {report.reference_fixes}")
    print(f"SKU reassignments: {report.sku_reassignments}")
    print(f"Parsed products: {len(validation.products)}")
    print(f"Parse issues: {len(validation.issues)}")
    if issue_counts:
        print("Issue breakdown:", dict(issue_counts))
    print(f"Duplicate SKUs remaining: {duplicate_codes}")
    return 0 if duplicate_codes == 0 and not issue_counts else 1


if __name__ == "__main__":
    raise SystemExit(main())
