from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import ExcelProduct, ParseIssue, ProductKind, WorkbookData
from .normalization import parse_price, parse_stock_status

HEADER_ALIASES = {
    "code": {"کد کالا", "code", "sku"},
    "reference": {"شناسه کالا", "reference", "parent", "parent sku", "parent_sku"},
    "title": {"نام کالا", "title", "name", "product name"},
    "attribute_name": {"تنوع", "attribute", "attribute name"},
    "attribute_value": {"مقدار تنوع", "attribute value", "option"},
    "price": {"قیمت", "price", "regular price", "regular_price"},
    "stock_marker": {"stock?", "stock", "stock status", "stock_status"},
}
REQUIRED_HEADERS = {"code", "reference", "title", "attribute_name", "attribute_value", "price"}


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _optional_column(values: tuple[Any, ...], field: str) -> int | None:
    for index, value in enumerate(values):
        header = _cell_text(value).casefold()
        aliases = HEADER_ALIASES.get(field, set())
        if header in {alias.casefold() for alias in aliases}:
            return index
    return None


def _read_stock_status(values: tuple[Any, ...], stock_index: int | None) -> str | None:
    if stock_index is None:
        return None
    raw = values[stock_index] if stock_index < len(values) else None
    if raw is None and stock_index >= len(values):
        return None
    return parse_stock_status(raw)


def _map_headers(values: tuple[Any, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, value in enumerate(values):
        header = _cell_text(value).casefold()
        for field, aliases in HEADER_ALIASES.items():
            if header in {alias.casefold() for alias in aliases}:
                mapping.setdefault(field, index)
                break
    missing = sorted(REQUIRED_HEADERS - mapping.keys())
    if missing:
        raise ValueError(f"Missing required Excel columns: {', '.join(missing)}")
    return mapping


def _has_variation_attrs(attribute_name: str, attribute_value: str) -> bool:
    return bool(attribute_name and attribute_value)


def read_products(path: str | Path) -> WorkbookData:
    """
    Parse the workbook without deduplicating SKUs.

    Rules (per row):
    - attribute name + value present  -> variation
    - code referenced by other rows   -> variable parent (when row has no attrs)
    - otherwise                       -> simple product

    The same SKU may appear on both a variation row and a parent row in the source
    file; both rows are kept because they represent different WooCommerce entities.
    """
    source = Path(path)
    if not source.is_file():
        raise FileNotFoundError(f"Excel file not found: {source}")

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0] if workbook.worksheets else None
        if worksheet is None:
            raise ValueError("Workbook has no worksheets")
        row_iterator = worksheet.iter_rows(values_only=True)
        try:
            headers = next(row_iterator)
        except StopIteration as exc:
            raise ValueError("Workbook is empty") from exc
        columns = _map_headers(headers)
        stock_index = _optional_column(headers, "stock_marker")

        raw_rows: list[dict[str, Any]] = []
        for row_number, values in enumerate(row_iterator, start=2):
            row = {field: values[index] if index < len(values) else None for field, index in columns.items()}
            if not any(_cell_text(value) for value in row.values()):
                continue
            row["row"] = row_number
            row["stock_status"] = _read_stock_status(values, stock_index)
            raw_rows.append(row)
    finally:
        workbook.close()

    codes = {_cell_text(row["code"]) for row in raw_rows if _cell_text(row["code"])}
    parent_codes = {
        _cell_text(row["reference"])
        for row in raw_rows
        if _cell_text(row["reference"])
        and _has_variation_attrs(_cell_text(row["attribute_name"]), _cell_text(row["attribute_value"]))
    }

    products: list[ExcelProduct] = []
    issues: list[ParseIssue] = []

    for row in raw_rows:
        row_number = int(row["row"])
        code = _cell_text(row["code"])
        reference = _cell_text(row["reference"])
        title = _cell_text(row["title"])
        attribute_name = _cell_text(row["attribute_name"])
        attribute_value = _cell_text(row["attribute_value"])

        def reject(reason: str) -> None:
            issues.append(ParseIssue(row_number, code, title, reason))

        if not code:
            reject("Missing product code")
            continue
        if not title:
            reject("Missing product title")
            continue
        if bool(attribute_name) != bool(attribute_value):
            reject("Variation attribute name and value must both be present")
            continue

        has_attrs = _has_variation_attrs(attribute_name, attribute_value)
        if has_attrs:
            if not reference:
                reject("Variation is missing its parent reference")
                continue
            if reference not in codes:
                reject(f'Variation parent "{reference}" is not present in the workbook')
                continue
            try:
                price = parse_price(row["price"])
            except ValueError as exc:
                reject(str(exc))
                continue
            if price is None:
                reject("Missing price")
                continue
            products.append(
                ExcelProduct(
                    row=row_number,
                    code=code,
                    reference=reference,
                    title=title,
                    attribute_name=attribute_name,
                    attribute_value=attribute_value,
                    price=price,
                    kind=ProductKind.VARIATION,
                    stock_status=row.get("stock_status"),
                )
            )
            continue

        if code in parent_codes:
            try:
                price = parse_price(row["price"])
            except ValueError as exc:
                reject(str(exc))
                continue
            products.append(
                ExcelProduct(
                    row=row_number,
                    code=code,
                    reference="",
                    title=title,
                    attribute_name="",
                    attribute_value="",
                    price=price,
                    kind=ProductKind.VARIABLE,
                    stock_status=None,
                )
            )
            continue

        try:
            price = parse_price(row["price"])
        except ValueError as exc:
            reject(str(exc))
            continue
        if price is None:
            reject("Missing price")
            continue
        products.append(
            ExcelProduct(
                row=row_number,
                code=code,
                reference=reference,
                title=title,
                attribute_name="",
                attribute_value="",
                price=price,
                kind=ProductKind.SIMPLE,
                stock_status=row.get("stock_status"),
            )
        )

    products.sort(key=lambda product: product.row)
    issues.sort(key=lambda issue: issue.row)
    return WorkbookData(products=products, issues=issues)
