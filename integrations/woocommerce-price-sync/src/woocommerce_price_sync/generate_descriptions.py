from __future__ import annotations

import argparse
import re
from pathlib import Path

from .description_product import (
    ProductDescription,
    ProductFacts,
    build_description,
    build_parser,
    default_paths,
    load_pending_products,
    merge_descriptions,
    read_catalog_products,
    read_descriptions_report,
    write_descriptions_report,
)
from .description_product_data import PRODUCT_FACTS, get_facts_for_product
from .web_research import (
    enrich_facts_from_research,
    load_research_cache,
    research_product_title,
    save_research_cache,
)


CATEGORY_HINTS: list[tuple[re.Pattern[str], str, str]] = [
    (re.compile(r"مولتی.?ویتامین|ویتامین|مکمل|کپسول|قرص|tablet|capsule", re.I), "مکمل غذایی", "قرص"),
    (re.compile(r"کرم|آبرسان|مرطوب|سرم|تونر|ماسک|لوسیون|ژل", re.I), "مراقبت پوست", "کرم"),
    (re.compile(r"شامپو|نرم.?کننده|ماسک مو", re.I), "مراقبت مو", "شامپو"),
    (re.compile(r"عطر|ادو|پرفیوم|perfume", re.I), "عطر", "ادو پرفیوم"),
    (re.compile(r"ریمل|خط چشم|پودر|کرم پودر|لیپ", re.I), "آرایشی", "محصول آرایشی"),
]


def detect_category(title: str) -> tuple[str, str]:
    for pattern, product_type, form in CATEGORY_HINTS:
        if pattern.search(title):
            return product_type, form
    return "مراقبت و زیبایی", "محصول"


def extract_brand(title: str) -> str:
    latin = re.findall(r"[A-Za-z][A-Za-z0-9&.'\- ]{1,30}", title)
    if latin:
        return latin[-1].strip()
    return ""


def _default_faq(title: str, product_type: str, brand: str) -> list[tuple[str, str]]:
    brand_text = f" برند {brand}" if brand else ""
    return [
        (f"روش استفاده از {title} چگونه است؟", "طبق دستور روی بسته‌بندی یا توصیه متخصص مصرف کنید."),
        (f"{title}{brand_text} برای چه منظوری است؟", f"این محصول در دسته {product_type} برای تکمیل روتین مراقبتی روزانه مناسب است."),
        (f"{title} برای چه افرادی مناسب است؟", "بسته به نوع پوست، مو یا نیاز تغذیه‌ای؛ جزئیات را در جدول مشخصات ببینید."),
        (f"نکات احتیاطی {title} چیست؟", "در صورت حساسیت، بارداری یا مصرف دارو با پزشک مشورت کنید."),
        (f"مزیت اصلی {title} چیست؟", "فرمول اختصاصی برای پشتیبانی از نیازهای روزانه مراقبت و زیبایی."),
        (f"آیا {title} اصل است؟", "برای خرید نسخه اصل با ضمانت اصالت، از فروشگاه آنلاین اسلی خرید کنید."),
    ]


def build_fallback_facts(title: str, code: str = "", woo_product_id: str | int | None = None) -> ProductFacts:
    product_type, form = detect_category(title)
    brand = extract_brand(title)
    volume_match = re.search(r"(\d+)\s*(میل|ml|گرم|g|عدد)", title, re.I)
    volume = volume_match.group(0) if volume_match else ""
    faq = _default_faq(title, product_type, brand)

    if product_type == "مکمل غذایی":
        return ProductFacts(
            title=title,
            code=code,
            woo_product_id=woo_product_id,
            brand=brand,
            product_type=product_type,
            form=form,
            volume=volume,
            summary=f"{title} مکملی حرفه‌ای برای پشتیبانی از نیازهای تغذیه‌ای روزانه است.",
            main_uses=["تأمین ریزمغذی‌ها", "حفظ انرژی روزانه", "پشتیبانی از سلامت عمومی"],
            benefits=[
                ("فرمول متعادل", "ترکیبی از ویتامین‌ها و مواد معدنی برای روتین روزانه."),
                ("مصرف آسان", "قابل استفاده در برنامه غذایی منظم."),
                ("کیفیت قابل اعتماد", "مناسب افرادی که به مکمل روزانه نیاز دارند."),
            ],
            ingredients=[
                ("ویتامین‌ها", "طبق فرمول محصول", "—"),
                ("مواد معدنی", "طبق فرمول محصول", "—"),
            ],
            usage_steps=[
                "طبق دستور روی بسته‌بندی یا توصیه متخصص مصرف کنید.",
                "همراه با آب کافی و ترجیحاً همراه غذا میل شود.",
                "از مصرف بیش از حد مجاز خودداری کنید.",
            ],
            suitable_for=["افراد بالغ", "کسانی با رژیم ناقص", "روتین سلامت روزانه"],
            cautions=["در بارداری و شیردهی با پزشک مشورت کنید.", "دور از دسترس کودکان نگهداری شود."],
            quick_facts=[
                ("نوع", product_type),
                ("فرم", form),
                ("برند", brand or "—"),
                ("حجم/تعداد", volume or "طبق بسته"),
            ],
            faq=faq,
        )

    if product_type == "عطر":
        return ProductFacts(
            title=title,
            code=code,
            woo_product_id=woo_product_id,
            brand=brand,
            product_type=product_type,
            form=form,
            volume=volume,
            summary=f"{title} عطری لوکس با رایحه‌ای ماندگار و شخصیتی متمایز است.",
            main_uses=["عطر روزانه", "مناسب مهمانی", "هدیه لوکس"],
            benefits=[
                ("ماندگاری مناسب", "رایحه‌ای پایدار برای استفاده روزانه یا شبانه."),
                ("شخصیت بویایی", "ترکیبی از نت‌های گرم و ظریف."),
                ("جلوه لوکس", "انتخابی مناسب برای علاقه‌مندان به عطرهای خاص."),
            ],
            ingredients=[
                ("نت آغازین", "ترکیبات معطر روشن", "—"),
                ("نت میانی", "گل‌ها و ادویه‌ها", "—"),
                ("نت پایه", "چوب و کهربا", "—"),
            ],
            usage_steps=[
                "روی نقاط نبض مانند مچ و گردن اسپری کنید.",
                "از مالش شدید پوست پس از اسپری خودداری کنید.",
                "برای ماندگاری بیشتر، روی پوست مرطوب استفاده کنید.",
            ],
            suitable_for=["بانوان و آقایان", "استفاده روزانه و مجلسی"],
            cautions=["از تماس مستقیم با چشم پرهیز کنید.", "دور از گرما و نور مستقیم نگهداری شود."],
            quick_facts=[
                ("نوع", "عطر"),
                ("برند", brand or "—"),
                ("غلظت", form),
            ],
            faq=faq,
        )

    return ProductFacts(
        title=title,
        code=code,
        woo_product_id=woo_product_id,
        brand=brand,
        product_type=product_type,
        form=form,
        volume=volume,
        summary=f"{title} انتخابی حرفه‌ای برای تکمیل روتین مراقبتی روزانه است.",
        main_uses=["مراقبت روزانه", "حفظ سلامت ظاهری", "تقویت روتین زیبایی"],
        benefits=[
            ("مراقبت موثر", "کمک به حفظ ظاهر سالم و متعادل."),
            ("بافت مناسب", "قابل استفاده در روتین صبح یا شب."),
            ("فرمول قابل اعتماد", "مناسب استفاده منظم."),
        ],
        ingredients=[
            ("ترکیبات فعال", "طبق فرمول محصول", "—"),
            ("مرطوب‌کننده‌ها", "پشتیبانی از رطوبت پوست", "—"),
        ],
        usage_steps=[
            "روی پوست تمیز و خشک استفاده کنید.",
            "مقدار مناسب را به‌صورت یکنواخت پخش کنید.",
            "صبح و شب در روتین مراقبتی تکرار کنید.",
        ],
        suitable_for=["انواع پوست با توجه به نوع محصول", "روتین روزانه"],
        cautions=["از تماس با چشم پرهیز کنید.", "در صورت حساسیت مصرف را متوقف کنید."],
        quick_facts=[
            ("نوع", product_type),
            ("فرم", form),
            ("برند", brand or "—"),
        ],
        faq=faq,
    )


def resolve_facts(
    title: str,
    code: str = "",
    woo_product_id: str | int | None = None,
    *,
    research_cache: dict | None = None,
    use_network: bool = True,
) -> ProductFacts:
    if title in PRODUCT_FACTS:
        facts = get_facts_for_product(title, code=code, woo_product_id=woo_product_id)
    else:
        facts = build_fallback_facts(title, code=code, woo_product_id=woo_product_id)

    if title not in PRODUCT_FACTS:
        research = research_product_title(
            title,
            cache=research_cache,
            use_network=use_network,
        )
        facts = enrich_facts_from_research(facts, research)
    return facts


def generate_description_for_product(
    title: str,
    code: str = "",
    woo_product_id: str | int | None = None,
    *,
    research_cache: dict | None = None,
    use_network: bool = True,
) -> ProductDescription:
    facts = resolve_facts(
        title,
        code=code,
        woo_product_id=woo_product_id,
        research_cache=research_cache,
        use_network=use_network,
    )
    return build_description(facts)


def run_batch(
    *,
    sync_report: Path | None = None,
    catalog: Path | None = None,
    output: Path | None = None,
    limit: int | None = 20,
    force: bool = False,
    catalog_only: bool = False,
    use_network: bool = True,
) -> Path:
    sync_path, catalog_path, output_path = default_paths()
    sync_report = sync_report or sync_path
    catalog = catalog or catalog_path
    output = output or output_path

    if force and catalog_only and catalog.is_file():
        products = read_catalog_products(catalog)
    else:
        products = load_pending_products(
            sync_report=sync_report,
            catalog=catalog,
            descriptions_report=output,
            limit=None if force else limit,
            force=force,
            catalog_only=catalog_only,
        )

    research_cache = load_research_cache()
    generated = [
        generate_description_for_product(
            product.title,
            code=product.code,
            woo_product_id=product.woo_product_id,
            research_cache=research_cache,
            use_network=use_network,
        )
        for product in products
    ]
    save_research_cache(research_cache)

    if force:
        merged = generated
    else:
        existing = read_descriptions_report(output) if output.is_file() else {}
        merged = merge_descriptions(existing, generated)

    input_path = catalog if catalog_only else (sync_report if sync_report.is_file() else catalog)
    return write_descriptions_report(
        output,
        merged,
        input_path=input_path,
    )


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)

    _, default_catalog, _ = default_paths()
    limit = None if args.all else args.limit
    report_path = run_batch(
        sync_report=args.sync_report,
        catalog=args.catalog or default_catalog,
        output=args.output,
        limit=limit,
        force=args.force,
        catalog_only=True,
        use_network=not args.no_web_research,
    )
    from openpyxl import load_workbook

    workbook = load_workbook(report_path, read_only=True, data_only=True)
    row_count = workbook.worksheets[0].max_row - 1
    workbook.close()
    print(f"Descriptions report written to: {report_path}")
    print(f"Total descriptions: {row_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
