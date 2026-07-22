from openpyxl import load_workbook

from woocommerce_price_sync.models import SyncResult
from woocommerce_price_sync.report import write_report


def test_writes_summary_and_row_level_audit(tmp_path):
    output = tmp_path / "report.xlsx"
    results = [
        SyncResult(
            excel_row=2,
            code="S-1",
            title="Product",
            product_type="simple",
            match_status="matched",
            action="would_update",
            woo_product_id=10,
            old_price="10",
            new_price="20",
            reason="Normalized title matched",
        ),
        SyncResult(
            excel_row=3,
            code="S-2",
            title="Missing",
            product_type="simple",
            match_status="not_found",
            action="would_create",
            new_price="30",
            reason="No match",
        ),
    ]

    write_report(output, results, apply=False, input_path="input.xlsx")

    workbook = load_workbook(output, read_only=True, data_only=True)
    assert workbook.sheetnames == ["Summary", "Results"]
    summary = dict(workbook["Summary"].iter_rows(values_only=True))
    assert summary["Mode"] == "dry-run"
    assert summary["Action: would_create"] == 1
    assert summary["Action: would_update"] == 1
    rows = list(workbook["Results"].iter_rows(values_only=True))
    assert rows[1][1] == "S-1"
    assert rows[1][7] == "would_update"
    workbook.close()
