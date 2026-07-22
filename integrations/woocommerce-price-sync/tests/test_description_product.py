from openpyxl import load_workbook

from woocommerce_price_sync.description_product import (
    ProductFacts,
    build_description,
    build_product_faq_list,
    build_product_table_rows,
    read_catalog_products,
    write_descriptions_report,
)
from woocommerce_price_sync.generate_descriptions import generate_description_for_product


def test_build_description_contains_sections(tmp_path):
    facts = ProductFacts(
        title="کرم تست",
        code="T1",
        summary="خلاصه تست.",
        main_uses=["آبرسانی"],
        benefits=[("نرمی", "کمک به نرمی پوست")],
        ingredients=[("گلیسیرین", "5%", "—")],
        usage_steps=["روی پوست تمیز استفاده کنید."],
        suitable_for=["پوست خشک"],
        cautions=["از چشم دور نگه دارید."],
        quick_facts=[("نوع", "کرم")],
        faq=[("چند بار؟", "روزانه یک بار.")],
    )
    result = build_description(facts)
    assert result.sku == "T1"
    assert "rose-table-block" in result.full_description
    assert "faq-lilac-block" in result.full_description
    assert "<h2>" in result.full_description
    assert result.short_description.count("<p>") == 2


def test_product_table_and_faq_use_product_details():
    facts = ProductFacts(
        title="آبرسان پوست خشک بایودرما",
        code="18636",
        brand="Bioderma",
        product_type="مراقبت پوست",
        form="کرم",
        volume="50ml",
        usage_steps=["صبح و شب روی پوست تمیز استفاده کنید."],
        suitable_for=["پوست خشک"],
        cautions=["از تماس با چشم پرهیز کنید."],
        benefits=[("آبرسانی", "رطوبت ماندگار")],
        ingredients=[("Aquagenium", "کمپلکس اختصاصی", "—")],
    )
    rows = build_product_table_rows(facts)
    faq = build_product_faq_list(facts)
    assert any("بایودرما" in row[0] or "Bioderma" in row[1] for row in rows)
    assert any("آبرسان پوست خشک بایودرما" in question for question, _ in faq)


def test_generate_description_for_product():
    item = generate_description_for_product(
        "A Z مولتی ویتامین بالای 50 سال بانوان",
        code="34376",
        web_search=False,
    )
    assert item.title.startswith("A Z")
    assert item.sku == "34376"
    assert "مولتی" in item.full_description


def test_read_catalog_products():
    from pathlib import Path

    catalog = Path(__file__).resolve().parents[1] / "data" / "catalog.xlsx"
    products = read_catalog_products(catalog)
    assert len(products) > 100


def test_write_descriptions_report(tmp_path):
    item = generate_description_for_product("کرم تست", code="T1", web_search=False)
    path = tmp_path / "descriptions-report.xlsx"
    write_descriptions_report(path, [item])
    assert path.is_file()
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    workbook.close()
    assert rows[0] == ("sku", "product_name", "short_description", "full_description")
    assert rows[1][0] == "T1"
