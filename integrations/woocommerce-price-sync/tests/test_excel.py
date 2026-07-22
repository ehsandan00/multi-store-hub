from openpyxl import Workbook

from woocommerce_price_sync.excel import read_products
from woocommerce_price_sync.models import ProductKind
from woocommerce_price_sync.normalization import normalize_text


def test_normalizes_persian_characters_and_zero_width():
    assert normalize_text("  مي\u200cشود  TEST ") == normalize_text("میشود test")
    assert normalize_text("كالا") == normalize_text("کالا")


def test_parses_and_classifies_persian_workbook(tmp_path):
    path = tmp_path / "products.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "کد کالا",
            "شناسه کالا",
            "نام کالا",
            "تنوع",
            "مقدار تنوع",
            "موجودی",
            "موجود می باشد",
            "قیمت",
        ]
    )
    sheet.append(["S-1", "P-S-1", "Simple", None, None, 2, "فعال", 12500])
    sheet.append(["P-1", None, "Variable", None, None, None, None, None])
    sheet.append(["V-1", "P-1", "Variable", "رنگ", "قرمز", 1, "فعال", 20000])
    sheet.append([None, "P-X", "Bad", None, None, None, None, 1])
    workbook.save(path)

    data = read_products(path)

    assert [product.kind for product in data.products] == [
        ProductKind.SIMPLE,
        ProductKind.VARIABLE,
        ProductKind.VARIATION,
    ]
    assert data.products[2].reference == "P-1"
    assert str(data.products[2].price) == "20000"
    assert len(data.issues) == 1
    assert data.issues[0].reason == "Missing product code"


def test_parses_stock_marker_column(tmp_path):
    path = tmp_path / "stock.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "کد کالا",
            "شناسه کالا",
            "نام کالا",
            "تنوع",
            "مقدار تنوع",
            "stock?",
            "قیمت",
        ]
    )
    sheet.append(["S-1", None, "In stock", None, None, 1, 100])
    sheet.append(["S-2", None, "Out of stock", None, None, 0, 200])
    sheet.append(["V-1", "P-1", "Variable", "رنگ", "قرمز", 0, 300])
    sheet.append(["P-1", None, "Variable", None, None, None, None])
    workbook.save(path)

    data = read_products(path)

    simple_in = next(product for product in data.products if product.code == "S-1")
    simple_out = next(product for product in data.products if product.code == "S-2")
    variation = next(product for product in data.products if product.code == "V-1")
    parent = next(product for product in data.products if product.code == "P-1")

    assert simple_in.stock_status == "instock"
    assert simple_out.stock_status == "outofstock"
    assert variation.stock_status == "outofstock"
    assert parent.stock_status is None


def test_ignores_inventory_quantity_column_without_stock_marker(tmp_path):
    path = tmp_path / "inventory.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "کد کالا",
            "شناسه کالا",
            "نام کالا",
            "تنوع",
            "مقدار تنوع",
            "موجودی",
            "قیمت",
        ]
    )
    sheet.append(["S-1", None, "Simple", None, None, 0, 100])
    workbook.save(path)

    data = read_products(path)

    assert data.products[0].stock_status is None


def test_keeps_same_sku_as_both_variation_and_parent_rows(tmp_path):
    path = tmp_path / "reuse.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "کد کالا",
            "شناسه کالا",
            "نام کالا",
            "تنوع",
            "مقدار تنوع",
            "موجودی",
            "موجود می باشد",
            "قیمت",
        ]
    )
    sheet.append(["P-ROOT", None, "Root", None, None, None, None, None])
    sheet.append(["PC-1", "P-ROOT", "Shared SKU", "نوع", "دکانت", 1, "فعال", 100])
    sheet.append(["PC-1", None, "Shared SKU parent", None, None, None, None, None])
    sheet.append(["PC-2", "PC-1", "Shared SKU parent", "نوع", "اورجینال", 1, "فعال", 200])
    workbook.save(path)

    data = read_products(path)

    kinds = [product.kind for product in data.products]
    assert kinds == [ProductKind.VARIABLE, ProductKind.VARIATION, ProductKind.VARIABLE, ProductKind.VARIATION]
    assert data.products[1].code == "PC-1"
    assert data.products[1].reference == "P-ROOT"
    assert data.products[2].code == "PC-1"
    assert data.products[3].reference == "PC-1"
    assert len(data.issues) == 0
